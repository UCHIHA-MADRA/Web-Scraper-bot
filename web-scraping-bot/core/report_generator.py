#!/usr/bin/env python3
"""
Advanced report generation with charts, analysis, error handling, and logging
"""
import pandas as pd
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns

# Import custom modules
from utils.logger import get_default_logger
from utils.exceptions import ReportGenerationError

class ReportGenerator:
    """
    Advanced report generator with error handling and logging
    """
    def __init__(self, config=None):
        """
        Initialize the report generator
        
        Args:
            config (dict, optional): Configuration for report generation
        """
        self.data = None
        self.config = config or {}
        self.logger = get_default_logger('report_generator')
        self.output_dir = self.config.get('output_dir', 'reports')
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Create subdirectories
        for subdir in ['daily', 'weekly', 'monthly', 'templates']:
            os.makedirs(os.path.join(self.output_dir, subdir), exist_ok=True)
    
    def load_data(self, data):
        """
        Load scraped data
        
        Args:
            data (list): List of dictionaries containing scraped data
        """
        try:
            self.logger.info(f"Loading data with {len(data)} records")
            self.data = pd.DataFrame(data)
            
            # Convert timestamp strings to datetime objects if needed
            if 'timestamp' in self.data.columns and isinstance(self.data['timestamp'].iloc[0], str):
                self.data['timestamp'] = pd.to_datetime(self.data['timestamp'])
                
            return True
        except Exception as e:
            self.logger.error(f"Error loading data: {e}")
            raise ReportGenerationError(f"Failed to load data: {str(e)}")
    
    def load_data_from_csv(self, filepath):
        """
        Load data from CSV file
        
        Args:
            filepath (str): Path to CSV file
        """
        try:
            self.logger.info(f"Loading data from CSV: {filepath}")
            self.data = pd.read_csv(filepath)
            
            # Convert timestamp strings to datetime objects
            if 'timestamp' in self.data.columns:
                self.data['timestamp'] = pd.to_datetime(self.data['timestamp'])
                
            return True
        except Exception as e:
            self.logger.error(f"Error loading data from CSV {filepath}: {e}")
            raise ReportGenerationError(f"Failed to load data from CSV: {str(e)}")
    
    def generate_price_trends(self):
        """
        Generate price trend analysis
        
        Returns:
            dict: Price trends by product
        """
        if self.data is None:
            self.logger.error("No data loaded for price trend analysis")
            return None
        
        try:
            self.logger.info("Generating price trend analysis")
            trends = {}
            
            # Ensure price column is numeric
            if 'price' in self.data.columns:
                # Convert price strings to numeric values
                self.data['price_numeric'] = self.data['price'].replace('[^\d.]', '', regex=True).astype(float)
            
                for product in self.data['product_name'].unique():
                    product_data = self.data[self.data['product_name'] == product]
                    
                    if len(product_data) > 1:
                        # Calculate statistics
                        avg_price = product_data['price_numeric'].mean()
                        min_price = product_data['price_numeric'].min()
                        max_price = product_data['price_numeric'].max()
                        
                        # Determine price trend
                        sorted_data = product_data.sort_values('timestamp')
                        if len(sorted_data) >= 2:
                            first_price = sorted_data['price_numeric'].iloc[0]
                            last_price = sorted_data['price_numeric'].iloc[-1]
                            
                            if last_price > first_price * 1.05:  # 5% increase
                                trend = 'increasing'
                            elif last_price < first_price * 0.95:  # 5% decrease
                                trend = 'decreasing'
                            else:
                                trend = 'stable'
                        else:
                            trend = 'insufficient_data'
                        
                        trends[product] = {
                            'avg_price': avg_price,
                            'min_price': min_price,
                            'max_price': max_price,
                            'price_change': trend
                        }
            
            return trends
        except Exception as e:
            self.logger.error(f"Error generating price trends: {e}")
            raise ReportGenerationError(f"Failed to generate price trends: {str(e)}")
    
    def generate_excel_report(self, filename=None, report_type='daily'):
        """
        Generate comprehensive Excel report with charts and formatting
        
        Args:
            filename (str, optional): Output filename
            report_type (str): Report type (daily, weekly, monthly)
            
        Returns:
            str: Path to generated report
        """
        if self.data is None or len(self.data) == 0:
            self.logger.error("No data available for Excel report generation")
            raise ReportGenerationError("No data available for report generation")
        
        try:
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{report_type}_report_{timestamp}.xlsx"
            
            # Determine output directory based on report type
            output_dir = os.path.join(self.output_dir, report_type)
            os.makedirs(output_dir, exist_ok=True)
            filepath = os.path.join(output_dir, filename)
            
            self.logger.info(f"Generating Excel report: {filepath}")
            
            # Create workbook and worksheets
            wb = Workbook()
            ws_summary = wb.active
            ws_summary.title = "Summary"
            ws_data = wb.create_sheet("Raw Data")
            ws_trends = wb.create_sheet("Price Trends")
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Raw Data Sheet
            self._generate_data_sheet(ws_data, header_font, header_fill, border)
            
            # Summary Sheet
            self._generate_summary_sheet(ws_summary, header_font, header_fill, border)
            
            # Price Trends Sheet
            self._generate_trends_sheet(ws_trends, header_font, header_fill, border)
            
            # Save workbook
            wb.save(filepath)
            self.logger.info(f"Excel report saved to {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating Excel report: {e}")
            raise ReportGenerationError(f"Failed to generate Excel report: {str(e)}")
    
    def _generate_data_sheet(self, ws, header_font, header_fill, border):
        """
        Generate the raw data sheet
        
        Args:
            ws (Worksheet): Worksheet to populate
            header_font (Font): Font for headers
            header_fill (PatternFill): Fill for headers
            border (Border): Border style
        """
        # Define headers based on available columns
        headers = self.data.columns.tolist()
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, (_, row_data) in enumerate(self.data.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                # Format timestamp if present
                if col_name == 'timestamp' and pd.notna(value):
                    if isinstance(value, pd.Timestamp):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_summary_sheet(self, ws, header_font, header_fill, border):
        """
        Generate the summary sheet
        
        Args:
            ws (Worksheet): Worksheet to populate
            header_font (Font): Font for headers
            header_fill (PatternFill): Fill for headers
            border (Border): Border style
        """
        # Title
        ws['A1'] = 'Competitor Analysis Report'
        ws['A1'].font = Font(size=16, bold=True)
        
        # Report metadata
        ws['A3'] = f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A4'] = f'Total products monitored: {self.data["product_name"].nunique()}'
        ws['A5'] = f'Total stores: {self.data["store"].nunique() if "store" in self.data.columns else "N/A"}'
        
        # Store summary
        if 'store' in self.data.columns:
            store_summary = {}
            for store in self.data['store'].unique():
                store_data = self.data[self.data['store'] == store]
                store_summary[store] = {
                    'products': store_data['product_name'].nunique(),
                    'errors': len(store_data[store_data['price'] == 'Error']) if 'price' in self.data.columns else 0
                }
            
            ws['A7'] = 'Store Summary:'
            ws['A7'].font = Font(bold=True)
            
            row = 8
            for store, stats in store_summary.items():
                ws[f'A{row}'] = f'  {store}: {stats["products"]} products, {stats["errors"]} errors'
                row += 1
        
        # Add a chart showing product distribution by store
        if 'store' in self.data.columns and len(self.data['store'].unique()) > 1:
            self._add_store_chart(ws)
    
    def _generate_trends_sheet(self, ws, header_font, header_fill, border):
        """
        Generate the price trends sheet
        
        Args:
            ws (Worksheet): Worksheet to populate
            header_font (Font): Font for headers
            header_fill (PatternFill): Fill for headers
            border (Border): Border style
        """
        # Generate price trends
        trends = self.generate_price_trends()
        
        if not trends:
            ws['A1'] = 'No price trend data available'
            return
        
        # Title
        ws['A1'] = 'Price Trend Analysis'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Product', 'Average Price', 'Minimum Price', 'Maximum Price', 'Price Trend']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        row = 4
        for product, data in trends.items():
            ws.cell(row=row, column=1, value=product).border = border
            ws.cell(row=row, column=2, value=data['avg_price']).border = border
            ws.cell(row=row, column=3, value=data['min_price']).border = border
            ws.cell(row=row, column=4, value=data['max_price']).border = border
            ws.cell(row=row, column=5, value=data['price_change']).border = border
            row += 1
        
        # Add a price comparison chart if we have multiple products
        if len(trends) > 1:
            self._add_price_chart(ws, row + 2, trends)
    
    def _add_store_chart(self, ws):
        """
        Add a chart showing product distribution by store
        
        Args:
            ws (Worksheet): Worksheet to add chart to
        """
        try:
            # Create a temporary data table for the chart
            row = 15
            ws.cell(row=row, column=1, value="Store")
            ws.cell(row=row, column=2, value="Products")
            
            store_counts = self.data.groupby('store')['product_name'].nunique()
            
            for i, (store, count) in enumerate(store_counts.items(), 1):
                ws.cell(row=row+i, column=1, value=store)
                ws.cell(row=row+i, column=2, value=count)
            
            # Create chart
            chart = BarChart()
            chart.title = "Products by Store"
            chart.style = 10
            chart.x_axis.title = "Store"
            chart.y_axis.title = "Number of Products"
            
            data = Reference(ws, min_col=2, min_row=row, max_row=row+len(store_counts))
            cats = Reference(ws, min_col=1, min_row=row+1, max_row=row+len(store_counts))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Add data labels
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            
            # Add chart to worksheet
            ws.add_chart(chart, "D15")
            
        except Exception as e:
            self.logger.warning(f"Failed to add store chart: {e}")
    
    def _add_price_chart(self, ws, start_row, trends):
        """
        Add a price comparison chart
        
        Args:
            ws (Worksheet): Worksheet to add chart to
            start_row (int): Starting row for chart data
            trends (dict): Price trend data
        """
        try:
            # Create a temporary data table for the chart
            ws.cell(row=start_row, column=1, value="Product")
            ws.cell(row=start_row, column=2, value="Average Price")
            
            for i, (product, data) in enumerate(trends.items(), 1):
                ws.cell(row=start_row+i, column=1, value=product)
                ws.cell(row=start_row+i, column=2, value=data['avg_price'])
            
            # Create chart
            chart = BarChart()
            chart.title = "Average Price Comparison"
            chart.style = 10
            chart.x_axis.title = "Product"
            chart.y_axis.title = "Price"
            
            data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row+len(trends))
            cats = Reference(ws, min_col=1, min_row=start_row+1, max_row=start_row+len(trends))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Add data labels
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            
            # Add chart to worksheet
            ws.add_chart(chart, "D" + str(start_row))
            
        except Exception as e:
            self.logger.warning(f"Failed to add price chart: {e}")
    
    def create_dashboard(self, output_path=None):
        """
        Create comprehensive dashboard using matplotlib/seaborn
        
        Args:
            output_path (str, optional): Path to save dashboard
            
        Returns:
            str: Path to generated dashboard
        """
        if self.data is None or len(self.data) == 0:
            self.logger.error("No data available for dashboard generation")
            return None
        
        try:
            if not output_path:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = os.path.join(self.output_dir, f"dashboard_{timestamp}.png")
            
            self.logger.info(f"Generating dashboard: {output_path}")
            
            # Set up the figure
            plt.figure(figsize=(12, 10))
            plt.suptitle("Web Scraping Analysis Dashboard", fontsize=16)
            
            # Set style
            sns.set_style("whitegrid")
            
            # Create subplots
            if 'price_numeric' in self.data.columns and 'product_name' in self.data.columns:
                # Price comparison
                plt.subplot(2, 2, 1)
                sns.barplot(x='product_name', y='price_numeric', data=self.data)
                plt.title("Price Comparison")
                plt.xticks(rotation=45, ha='right')
                plt.tight_layout()
            
            # Save the dashboard
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            plt.close()
            
            self.logger.info(f"Dashboard saved to {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"Error creating dashboard: {e}")
            return None
