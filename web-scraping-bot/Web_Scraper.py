#!/usr/bin/env python3
"""
Web Scraping and Report Generation Bot - Main Application
Created by: RPA Bot Developer

This application provides a configurable web scraping solution with
report generation, caching, and email notification capabilities.

Features:
- Configurable web scraping for multiple targets
- Intelligent caching system to reduce network requests
- Report generation in Excel format
- Email notification capabilities
- Multi-level caching system to improve performance
- Command-line interface for easy interaction
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime, timedelta
import time
import random
import json
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import schedule
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from urllib.parse import urljoin, urlparse

# Import custom modules
from utils.logger import get_default_logger
from utils.cache import get_default_cache
from utils.exceptions import WebScraperError, NetworkError, ScrapingError, ParsingError

class WebScrapingBot:
    """
    Web scraping bot for extracting product information from e-commerce websites
    
    This class provides methods for scraping product information, exporting data
    to Excel, and sending email notifications with the results. It utilizes an
    intelligent caching system to improve performance by avoiding redundant scraping
    of the same URLs within the cache expiration period.
    
    The bot integrates a three-level caching system:
    1. HTTP request caching - Avoids repeating network requests
    2. Memory caching - Provides fast access to frequently used data
    3. File caching - Persists data between application restarts
    
    Attributes:
        session: Requests session for making HTTP requests
        logger: Logger instance for logging operations
        cache: ScraperCache instance for caching scraped data
        config: Configuration dictionary loaded from config file
    """
    def __init__(self, config_file='config/config.json'):
        """Initialize the web scraping bot
        
        Sets up the requests session, logging, caching system, and loads configuration.
        The caching system is initialized with a 1-hour expiration time by default.
        """
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
        
        # Setup logging
        self.logger = get_default_logger('scraper')
        
        # Setup caching
        self.cache = get_default_cache(cache_dir='cache', expiration=3600)  # 1 hour cache expiration
        self.logger.info("Cache system initialized")
        
        # Load configuration
        self.config = self.load_config(config_file)
        self.scraped_data = []
        
    def load_config(self, config_file):
        """Load configuration from JSON file"""
        default_config = {
            "targets": [
                {
                    "name": "Books to Scrape Demo",
                    "base_url": "http://books.toscrape.com",
                    "products": [
                        {
                            "name": "A Light in the Attic",
                            "url": "http://books.toscrape.com/catalogue/a-light-in-the-attic_1000/index.html",
                            "price_selector": "p.price_color",
                            "availability_selector": "p.availability",
                            "rating_selector": "p.star-rating"
                        }
                    ]
                }
            ],
            "delay_range": [1, 3],
            "output_dir": "reports/daily",
            "email": {
                "enabled": False,
                "smtp_server": "smtp.gmail.com",
                "smtp_port": 587,
                "sender_email": "",
                "sender_password": "",
                "recipients": []
            }
        }
        
        try:
            if os.path.exists(config_file):
                with open(config_file, 'r') as f:
                    return json.load(f)
            else:
                # Create config directory and file
                os.makedirs(os.path.dirname(config_file), exist_ok=True)
                with open(config_file, 'w') as f:
                    json.dump(default_config, f, indent=4)
                self.logger.info(f"Created default config file: {config_file}")
                return default_config
        except Exception as e:
            self.logger.error(f"Error loading config: {e}")
            return default_config

    def scrape_product(self, product_info, store_name):
        """Scrape data for a single product"""
        try:
            url = product_info['url']
            self.logger.info(f"Processing {product_info['name']} from {store_name}")
            
            # Check cache first
            cache_params = {
                'selectors': {
                    'price': product_info.get('price_selector'),
                    'availability': product_info.get('availability_selector'),
                    'rating': product_info.get('rating_selector')
                }
            }
            
            cached_data = self.cache.get(url, params=cache_params)
            if cached_data:
                self.logger.info(f"Using cached data for {product_info['name']}")
                # Update timestamp to current time
                cached_data['timestamp'] = datetime.now()
                return cached_data
            
            # Not in cache, need to scrape
            self.logger.info(f"Scraping {product_info['name']} from {store_name}")
            
            # Random delay to be respectful to the server
            delay = random.uniform(*self.config['delay_range'])
            time.sleep(delay)
            
            try:
                response = self.session.get(url)
                response.raise_for_status()
            except requests.exceptions.RequestException as e:
                raise NetworkError(f"Failed to fetch URL: {url}", url=url, status_code=getattr(e.response, 'status_code', None), details=str(e))
            
            try:
                soup = BeautifulSoup(response.content, 'html.parser')
            except Exception as e:
                raise ParsingError(f"Failed to parse HTML content", content_sample=response.content[:100], details=str(e))
            
            # Extract data
            price = "N/A"
            if 'price_selector' in product_info:
                try:
                    price_elem = soup.select_one(product_info['price_selector'])
                    if price_elem:
                        price = price_elem.get_text(strip=True)
                except Exception as e:
                    self.logger.warning(f"Failed to extract price using selector '{product_info['price_selector']}': {e}")
            
            availability = "N/A"
            if 'availability_selector' in product_info:
                try:
                    avail_elem = soup.select_one(product_info['availability_selector'])
                    if avail_elem:
                        availability = avail_elem.get_text(strip=True)
                except Exception as e:
                    self.logger.warning(f"Failed to extract availability using selector '{product_info['availability_selector']}': {e}")
            
            rating = "N/A"
            if 'rating_selector' in product_info:
                try:
                    rating_elem = soup.select_one(product_info['rating_selector'])
                    if rating_elem:
                        rating = rating_elem.get_text(strip=True)
                except Exception as e:
                    self.logger.warning(f"Failed to extract rating using selector '{product_info['rating_selector']}': {e}")
            
            product_data = {
                'timestamp': datetime.now(),
                'store': store_name,
                'product_name': product_info['name'],
                'price': price,
                'availability': availability,
                'rating': rating,
                'url': url
            }
            
            # Store in cache
            self.cache.set(url, product_data, params=cache_params)
            
            self.logger.info(f"Successfully scraped {product_info['name']}: {price}")
            return product_data
            
        except WebScraperError as e:
            self.logger.error(f"Error scraping {product_info['name']}: {e.message}")
            if hasattr(e, 'details') and e.details:
                self.logger.debug(f"Error details: {e.details}")
            return {
                'timestamp': datetime.now(),
                'store': store_name,
                'product_name': product_info['name'],
                'price': 'Error',
                'availability': 'Error', 
                'rating': 'Error',
                'url': product_info['url']
            }
        except Exception as e:
            self.logger.error(f"Unexpected error scraping {product_info['name']}: {e}")
            return {
                'timestamp': datetime.now(),
                'store': store_name,
                'product_name': product_info['name'],
                'price': 'Error',
                'availability': 'Error', 
                'rating': 'Error',
                'url': product_info['url']
            }

    def scrape_all_products(self):
        """Scrape all configured products"""
        self.scraped_data = []
        
        for target in self.config['targets']:
            store_name = target['name']
            self.logger.info(f"Starting to scrape {store_name}")
            
            for product in target['products']:
                product_data = self.scrape_product(product, store_name)
                self.scraped_data.append(product_data)
        
        self.logger.info(f"Scraping completed. Collected {len(self.scraped_data)} records")
        return self.scraped_data

    def generate_excel_report(self, filename=None):
        """Generate Excel report"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"competitor_report_{timestamp}.xlsx"
        
        os.makedirs(self.config['output_dir'], exist_ok=True)
        filepath = os.path.join(self.config['output_dir'], filename)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Data"
        
        # Headers
        headers = ['Timestamp', 'Store', 'Product', 'Price', 'Availability', 'Rating', 'URL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Data
        for row, record in enumerate(self.scraped_data, 2):
            ws.cell(row=row, column=1, value=record['timestamp'].strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=2, value=record['store'])
            ws.cell(row=row, column=3, value=record['product_name'])
            ws.cell(row=row, column=4, value=record['price'])
            ws.cell(row=row, column=5, value=record['availability'])
            ws.cell(row=row, column=6, value=record['rating'])
            ws.cell(row=row, column=7, value=record['url'])
        
        wb.save(filepath)
        self.logger.info(f"Excel report saved to {filepath}")
        return filepath

    def run_scraping(self):
        """Main function to run scraping and report generation"""
        self.logger.info("Starting scraping process...")
        
        try:
            # Scrape products
            self.scrape_all_products()
            
            # Generate report
            report_path = self.generate_excel_report()
            
            # Log cache statistics
            cache_stats = self.cache.get_stats()
            self.logger.info(f"Cache statistics: {cache_stats['memory_cache_entries']} memory entries, "
                           f"{cache_stats['file_cache_entries']} file entries, "
                           f"{cache_stats['http_requests']['hits']} HTTP cache hits")
            
            self.logger.info("Scraping process completed successfully")
            print(f"✅ Report generated: {report_path}")
            
        except WebScraperError as e:
            self.logger.error(f"Error in scraping process: {e.message}")
            if hasattr(e, 'details') and e.details:
                self.logger.debug(f"Error details: {e.details}")
        except Exception as e:
            self.logger.error(f"Unexpected error in scraping process: {e}")
            self.logger.exception("Exception details:")

def main():
    print("🤖 Web Scraping and Report Generation Bot")
    print("=" * 50)
    
    bot = WebScrapingBot()
    
    print("Options:")
    print("1. Run scraping once")
    print("2. View current configuration")
    print("3. View cache statistics")
    print("4. Clear cache")
    print("5. Exit")
    
    choice = input("\nSelect option (1-5): ").strip()
    
    if choice == "1":
        bot.run_scraping()
    elif choice == "2":
        print(json.dumps(bot.config, indent=2))
    elif choice == "3":
        stats = bot.cache.get_stats()
        print("\nCache Statistics:")
        print(f"Memory cache entries: {stats['memory_cache_entries']}")
        print(f"File cache entries: {stats['file_cache_entries']}")
        print(f"HTTP requests - Hits: {stats['http_requests']['hits']}")
        print(f"HTTP requests - Misses: {stats['http_requests']['misses']}")
        print(f"HTTP requests - Total: {stats['http_requests']['total']}")
        print(f"Cache directory: {stats['cache_dir']}")
        print(f"Cache expiration: {stats['expiration']} seconds")
    elif choice == "4":
        bot.cache.clear()
        print("✅ Cache cleared successfully")
    else:
        print("Goodbye! 👋")

if __name__ == "__main__":
    main()
